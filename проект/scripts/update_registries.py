# -*- coding: utf-8 -*-
"""Добавление новых источников в реестры доклада (S023-S027, M-comf_*, M-hydr_*)."""
import openpyxl, shutil, os

BASE = r'C:\Users\Evgenii\OneDrive\Документы\DSeek\repo-svf'

def load(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = []
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            hdr = list(row)
        else:
            rows.append(list(row))
    return wb, ws, hdr, rows

# ---------- SOURCE_REGISTRY ----------
p = os.path.join(BASE, 'проект', 'SOURCE_REGISTRY.xlsx')
wb, ws, hdr, rows = load(p)

# обновляем S003 (СанПиН): добавляем 4282-26
for r in rows:
    if r and r[0] == 'S003':
        r[1] = ('СанПиН 2.3/2.4.3590-20 (пост. ГГСВ РФ от 27.10.2020 № 32), п. 7.1.11–7.1.14 — до 01.09.2026; '
                'СанПиН 2.3/2.4.4282-26 (пост. ГГСВ РФ от 02.06.2026 № 18, зарег. в Минюсте 02.06.2026 № 86854), '
                'п. 56 подп. 11–14, п. 62, раздел VI')
        r[4] = 'https://rovenkiadm.gosuslugi.ru/netcat_files/userfiles/SanPiN_2.4.4282-26.pdf'
        r[6] = '3590-20 действует до 01.09.2026; 4282-26 действует с 01.09.2026 до 01.09.2032'
        r[7] = ('прямые нормы о питании в СОСО: не менее 3 раз/день, диетическое по показаниям (п. 56 подп. 11); '
                'суточные пробы при аутсорсинге (подп. 12), бракераж (подп. 13), привлекаемые предприятия (подп. 14); '
                'питьевой режим (п. 62); кейтеринг (раздел VI)')
        r[10] = 'текст 4282-26 сверен по официальной публикации 17.08.2026'
        break

new_s = [
 ['S023','ФЗ от 05.04.2013 № 44-ФЗ «О контрактной системе в сфере закупок товаров, работ, услуг для обеспечения государственных и муниципальных нужд»','федеральный закон / A','Госдума','https://www.consultant.ru/document/cons_doc_LAW_144624/','23.07.2026','действует','контрактная система закупок; правовая рамка аутсорсинга питания','глава 44, раздел 30','высокая','—'],
 ['S024','Лозовская С.О. «Правосубъектность лиц, признанных недееспособными и ограниченно дееспособными вследствие психического расстройства»','научная статья / B','Вестник Университета имени О.Е. Кутафина (МГЮА), 2023, № 5, с. 139–146','https://vestnik.msal.ru/jour/article/view/2058','23.07.2026','опубликовано','практика применения п. 2 ст. 30 ГК: институт невостребован, критерии размыты','раздел 16','средняя','DOI 10.17803/2311-5998.2023.105.5.139-146'],
 ['S025','«Кейтеринг в организациях социального обслуживания» (отраслевое издание, 2022)','отраслевая статья / C','Профиздат','https://www.profiz.ru/sec/4_2022/Kejtering/','23.07.2026','опубликовано','практика передачи питания на аутсорсинг в СОСО; вариативность через контракт','глава 44','средняя','—'],
 ['S026','«Услуга приготовления диетических блюд» — справочные материалы по организации лечебного питания','отраслевой материал / C','Национальная ассоциация клинического питания (НАКП)','https://nakp.org/usluga-prigotovleniya-dieticheskikh-blyud/','23.07.2026','опубликовано','опыт привлечённых предприятий для диетического питания','глава 44','средняя','—'],
 ['S027','СП 2.4.3648-20 «Санитарно-эпидемиологические требования к организациям воспитания и обучения, отдыха и оздоровления детей и молодежи» (пост. ГГСВ РФ от 28.09.2020 № 28, зарег. в Минюсте 18.12.2020 № 61573)','санитарные правила / A','Роспотребнадзор','https://www.consultant.ru/document/cons_doc_LAW_371594/','23.07.2026','действует (ред. от 24.12.2025)','общие санитарные требования к детским организациям','приложение 41','высокая','—'],
]
for r in new_s:
    rows.append(r)

# перезапись листа
ws.delete_rows(1, ws.max_row)
ws.append(hdr)
for r in rows:
    ws.append(r)
wb.save(p)
print('SOURCE_REGISTRY: rows =', len(rows))

# ---------- MEDICAL_EVIDENCE ----------
p2 = os.path.join(BASE, 'проект', 'MEDICAL_EVIDENCE.xlsx')
wb2, ws2, hdr2, rows2 = load(p2)
new_m = [
 ['M-comf_oralfeed','Oral feeding options for people with dementia: a systematic review','систематический обзор (2011); PMID 21391936','https://pubmed.ncbi.nlm.nih.gov/21391936/','C','верифицировано 17.08.2026','раздел 12 (comfort feeding)'],
 ['M-comf_jnha','Comfort feeding in hospitalised people with dementia: a retrospective study of survival following comfort feeding recommendations','J Nutr Health Aging, 2024; DOI 10.1016/j.jnha.2024.100362','https://doi.org/10.1016/j.jnha.2024.100362','C','верифицировано 17.08.2026','раздел 12 (comfort feeding, выживаемость)'],
 ['M-hydr_esp','ESPEN guideline on clinical nutrition and hydration in geriatrics','ESPEN (Европейское общество клинического питания и метаболизма); Clin Nutr','https://www.espen.org/files/ESPEN-Guidelines/ESPEN_guideline_on_clincal_nutrition_and_hydration_in_geriatrics.pdf','A','верифицировано 17.08.2026','приложение 40 (питьевой режим)'],
 ['M-hydr_palt','Dehydration and Fluid Maintenance in the Long-Term Care Setting — клиническое руководство','PALTmed (Общество паллиативной медицины), 2024','https://paltmed.org/sites/default/files/2024-10/Dehydration%20CPG%2008092024.pdf','B','верифицировано 17.08.2026','приложение 40 (питьевой режим в LTC)'],
 ['M-hydr_low','Narrative Review of Low-Intake Dehydration in Older Adults','обзор (2021); PMC8470893','https://pmc.ncbi.nlm.nih.gov/articles/PMC8470893/','C','верифицировано 17.08.2026','приложение 40 (распространённость дегидратации)'],
]
for r in new_m:
    rows2.append(r)
ws2.delete_rows(1, ws2.max_row)
ws2.append(hdr2)
for r in rows2:
    ws2.append(r)
wb2.save(p2)
print('MEDICAL_EVIDENCE: rows =', len(rows2))

# ---------- копии реестров в 04_приложения_и_реестры и семинар_пакет ----------
for dst in [os.path.join(BASE, '04_приложения_и_реестры'), os.path.join(BASE, 'семинар_пакет', '04_приложения_и_реестры')]:
    shutil.copy2(p, os.path.join(dst, 'SOURCE_REGISTRY.xlsx'))
    shutil.copy2(p2, os.path.join(dst, 'MEDICAL_EVIDENCE.xlsx'))
    print('copied to', dst)
