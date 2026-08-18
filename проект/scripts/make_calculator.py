# -*- coding: utf-8 -*-
"""Калькулятор стоимости вариативности (XLSX с формулами)."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
head_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
head_font = Font(bold=True, color='FFFFFF')

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def input_cell(ws, row, col):
    ws.cell(row=row, column=col).fill = yellow

# --- Лист 1: Пары блюд ---
ws = wb.active
ws.title = 'Пары блюд'
ws.append(['Позиция', 'Вариант А', 'Вариант Б', 'Норма закладки А, г/порц.', 'Норма закладки Б, г/порц.', 'Цена А, руб/кг', 'Цена Б, руб/кг', 'Стоимость порции А, руб', 'Стоимость порции Б, руб', 'Разница, руб/порц.', 'Порций в месяц', 'Разница, руб/мес'])
style_header(ws, 12)
rows = [
    ['Каша гарнирная', 'Гречка', 'Рис', 200, 200, 90, 110, None, None, None, 600, None],
    ['Белковая позиция', 'Котлета мясная', 'Котлета рыбная', 100, 100, 320, 260, None, None, None, 600, None],
    ['Первое', 'Щи', 'Борщ', 250, 250, 60, 65, None, None, None, 600, None],
    ['Напиток', 'Компот', 'Морс', 200, 200, 50, 55, None, None, None, 600, None],
]
for r in rows:
    ws.append(r)
for i in range(2, 2 + len(rows)):
    input_cell(ws, i, 4); input_cell(ws, i, 5); input_cell(ws, i, 6); input_cell(ws, i, 7); input_cell(ws, i, 11)
    ws.cell(row=i, column=8).value = f'=ROUND(C{i}*D{i}/1000*F{i},2)'
    ws.cell(row=i, column=9).value = f'=ROUND(C{i}*E{i}/1000*G{i},2)'
    ws.cell(row=i, column=10).value = f'=ROUND(H{i}-I{i},2)'
    ws.cell(row=i, column=12).value = f'=ROUND(K{i}*J{i},2)'
ws.append([])
ws.append(['Итого разница, руб/мес', None, None, None, None, None, None, None, None, None, None, f'=SUM(L2:L{1+len(rows)})'])
input_cell(ws, 2 + len(rows), 1)

# --- Лист 2: Отходы ---
ws2 = wb.create_sheet('Отходы')
ws2.append(['Показатель', 'Значение', 'Ед.', 'Комментарий'])
style_header(ws2, 4)
data2 = [
    ['Число жителей', 150, 'чел.', 'входные данные учреждения'],
    ['Тарелочные остатки', 80, 'г/порцию', 'замер 7 дней (приложение 25)'],
    ['Приёмов пищи в день', 3, '', ''],
    ['Средняя закупочная цена сырья', 250, 'руб/кг', 'по фактуре'],
    ['Отходы в день', None, 'кг', ''],
    ['Строка «выброшенных денег» в день', None, 'руб', ''],
    ['Строка «выброшенных денег» в год', None, 'руб', '365 дней'],
]
for r in data2:
    ws2.append(r)
for i in range(2, 6):
    input_cell(ws2, i, 2)
ws2.cell(row=6, column=2).value = '=ROUND(B2*B3*B4/1000,1)'
ws2.cell(row=7, column=2).value = '=ROUND(B6*B5,0)'
ws2.cell(row=8, column=2).value = '=ROUND(B7*365,0)'

# --- Лист 3: Сценарии ---
ws3 = wb.create_sheet('Сценарии')
ws3.append(['Сценарий снижения отходов', 'Снижение, %', 'Экономия в день, руб', 'Экономия в год, руб', 'Пометка'])
style_header(ws3, 5)
for i, pct in enumerate([10, 20, 30]):
    ws3.append([f'Снижение на {pct}%', pct, None, None, 'гипотеза — подтверждается только повторным замером через 90 дней'])
    r = 2 + i
    input_cell(ws3, r, 2)
    ws3.cell(row=r, column=3).value = f"=ROUND('Отходы'!B7*B{r}/100,0)"
    ws3.cell(row=r, column=4).value = f"=ROUND('Отходы'!B8*B{r}/100,0)"

wb.save(str(Path(__file__).resolve().parents[1] / 'Калькулятор_стоимости_вариативности.xlsx'))
print('calculator saved')
